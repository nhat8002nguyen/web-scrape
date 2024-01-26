import os
from time import sleep
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import shutil
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from seleniumbase import Driver
from json import loads

from dotenv import load_dotenv
load_dotenv()

PROJECT_PATH = os.environ["PROJECT_PATH"]
COVER_IMAGES_PATH = PROJECT_PATH + "/cover_images"

# Define the headers for HTTP requests to look like a browser request
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# Define the base URL for Google Search
GOOGLE_SEARCH_URL = 'https://www.google.com/search'

# Create directory for cover images if it doesn't exist
if not os.path.exists(COVER_IMAGES_PATH):
    os.makedirs(COVER_IMAGES_PATH)


def download_image(url, filename):
    '''Function to download and save an image'''
    response = requests.get(url, stream=True)
    with open(filename, 'wb') as file:
        shutil.copyfileobj(response.raw, file)


def get_cover_image_url(session, amazon_url: str):
    '''Add your code to extract the image URL'''
    resp = session.get(amazon_url, headers=HEADERS)

    soup = BeautifulSoup(resp.content, 'html.parser')

    img = soup.select_one("#imgTagWrapperId > img")
    if img == None:
        return None

    img_src = img.get("src")

    return img_src


def search_google(book_name, author_name):
    '''Function to search Google for Amazon book page'''
    # Construct search query
    query = f'site:amazon.com "{book_name}" "{author_name}"'

    # Make request to Google
    google_response = requests.get(
        GOOGLE_SEARCH_URL, headers=HEADERS, params={'q': query})

    soup = BeautifulSoup(google_response.content, 'html.parser')

    # parse Google search results, extract Amazon links
    first_result = soup.select_one("div.MjjYud .yuRUbf a")
    if first_result == None:
        return None

    link = first_result.get("href")
    return link


# Load the Excel workbook and get the first sheet
wb = load_workbook(f'{PROJECT_PATH}/27-categories-books.xlsx')
sheet = wb.active

print("Started the program!")
driver: WebDriver = Driver(uc=True, no_sandbox=True, headless=True)
wait = WebDriverWait(driver, 30)

session = requests.Session()
with open("./cookies.json", "r") as file:
    cookies = list(loads(file.read()))

    for cookie in cookies:
        session.cookies.set(
            name=cookie['name'],
            value=cookie['value'],
            domain=cookie.get('domain'),
            path=cookie.get('path'),
            secure=cookie.get('secure'),
            # Add additional attrs as necessary
            rest={'HttpOnly': cookie.get('httpOnly')}
        )

# Loop over the rows in the Excel sheet
count = 0
for row in sheet.iter_rows(min_row=2):  # assuming first row is the header
    count += 1
    if count > 50:
        break

    # replace 0 with actual index for book name
    book_name = row[1].value.replace("/", "-")
    # replace 1 with actual index for author name
    author_name = row[0].value.replace("/", "-")

    # Google Search to find Amazon book page
    amazon_url = search_google(book_name, author_name)

    if amazon_url == None:
        continue

    # Get cover image URL from Amazon book page
    image_url = get_cover_image_url(session, amazon_url)
    if image_url == None:
        continue

    # Download and save the cover image
    # Consider sanitizing filename
    image_filename = os.path.join(
        COVER_IMAGES_PATH, f'{book_name}_{author_name}.jpg')
    download_image(image_url, image_filename)

    sleep(3)

driver.close()
