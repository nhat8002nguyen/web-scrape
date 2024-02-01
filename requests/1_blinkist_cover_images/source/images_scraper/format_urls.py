from openpyxl import load_workbook

# Define the path to your Excel file
excel_path = "/Users/nhatnguyen/Workspaces/web-scrape/requests/1_blinkist_cover_images/27-categories-books.xlsx"

# Load the workbook and select the active worksheet
workbook = load_workbook(excel_path)
sheet = workbook.active

# Assuming URLs are in the first column (column A)
for row in sheet.iter_rows(min_row=2, max_col=1):
    for cell in row:
        if cell.value:
            # Replace spaces in the URL with '%20' or remove them entirely
            cell.value = cell.value.replace(' ', '%20')

# Save the workbook
workbook.save(excel_path)
print("All URLs have been formatted correctly.")
