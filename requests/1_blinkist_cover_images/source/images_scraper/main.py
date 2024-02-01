import os
from time import sleep
import traceback
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import shutil
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from seleniumbase import Driver
from json import loads
import argparse

from dotenv import load_dotenv
load_dotenv()

PROJECT_PATH = os.environ["PROJECT_ROOT"]
COVER_IMAGES_PATH = PROJECT_PATH + "/cover_images"

# Define the headers for HTTP requests to look like a browser request
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# Define the base URL for Google Search
GOOGLE_SEARCH_URL = 'https://www.google.com/search'

AMAZON_SEARCH_URL = 'https://www.amazon.com/s'

# Create directory for cover images if it doesn't exist
if not os.path.exists(COVER_IMAGES_PATH):
    os.makedirs(COVER_IMAGES_PATH)


def download_image(url, filename):
    '''Function to download and save an image'''
    response = requests.get(url, stream=True)
    with open(filename, 'wb') as file:
        shutil.copyfileobj(response.raw, file)


def get_cover_image_url(session, amazon_url: str):
    '''Add your code to extract the image URL'''
    img_src = None
    try:
        count = 0
        while count < 20:
            count += 1
            resp = session.get(amazon_url, headers=HEADERS)

            if resp.status_code == 200:
                soup = BeautifulSoup(resp.content, 'html.parser')

                img = soup.select_one("#imgTagWrapperId > img")
                if img == None:
                    return None

                img_src = img.get("src")
                return img_src
            else:
                session = requests.Session()
                session.proxies = {
                    'http': "http://proxy005844-rotate:proxy005844@p.webshare.io:80",
                    'https': "http://proxy005844-rotate:proxy005844@p.webshare.io:80",
                }
                if count < 10:
                    print(f"Fail with status {resp.status_code}, retrying...")
    except:
        print(f"Fail with {amazon_url}")

    return img_src


def search_google(book_name, author_name):
    '''Function to search Google for Amazon book page'''
    try:
        count = 0
        query = f'site:amazon.com "{book_name}" "{author_name}"'
        while count < 20:
            count += 1
            response = requests.get(
                GOOGLE_SEARCH_URL, headers=HEADERS, params={'q': query}, proxies={
                    'http': "http://proxy005844-rotate:proxy005844@p.webshare.io:80",
                    'https': "http://proxy005844-rotate:proxy005844@p.webshare.io:80",
                })
            # Check if the request was successful
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')

                # parse Google search results, extract Amazon links
                first_result = soup.select_one("div.MjjYud .yuRUbf a")
                if first_result == None:
                    return None

                link = first_result.get("href")
                return link
            else:
                query = f'site:amazon.com {book_name} {author_name}'
                # If the response status code is not 200, log the error and try the next proxy
                print(
                    f"Proxy failed with status code {response.status_code}.")
    except:
        # If an error occurs, print the traceback and try the next proxy
        print(f"Proxy threw an exception.")
        traceback.print_exc()


def create_image_name(sheet_index: int, book_name: str, author_name: str) -> str:
    book_name = book_name.replace("/", "-")
    author_name = author_name.replace("/", "-")

    return f"{sheet_index}_{book_name}_{author_name}"


def main():
    # Create the parser
    parser = argparse.ArgumentParser(description='Process some integers.')

    # Add arguments
    parser.add_argument(
        '--start', dest='start_index', type=int, required=True,
        help='start is the start index of rows in the sheet.'
    )

    parser.add_argument(
        '--end', dest='end_index', type=int, required=True,
        help='end is the end index of rows in the sheet.'
    )

    # Parse arguments
    args = parser.parse_args()

    if args.start_index > args.end_index:
        return print("Start index should be less than end index.")

    min_row = args.start_index+1
    max_row = args.end_index+1

    # Load the Excel workbook and get the first sheet
    wb = load_workbook(f'{PROJECT_PATH}/27-categories-books.xlsx')
    sheet = wb.active

    session = requests.Session()
    with open(f"{PROJECT_PATH}/cookies.json", "r") as file:
        cookies = list(loads(file.read()))

        for cookie in cookies:
            session.cookies.set(
                name=cookie['name'],
                value=cookie['value'],
                domain=cookie.get('domain'),
                path=cookie.get('path'),
                secure=cookie.get('secure'),
                # Add additional attrs as necessary
                rest={'HttpOnly': cookie.get('httpOnly')}
            )

    # Loop over the rows in the Excel sheet
    # assuming first row is the header
    for i, row in enumerate(sheet.iter_rows(min_row=min_row, max_row=max_row)):
        image_url = row[0].value
        if image_url:
            continue

        # replace 0 with actual index for book name
        book_name = row[2].value
        # replace 1 with actual index for author name
        author_name = row[1].value

        # Google Search to find Amazon book page
        amazon_url = search_google(book_name, author_name)

        if amazon_url == None:
            continue

        # Get cover image URL from Amazon book page
        image_url = get_cover_image_url(session, amazon_url)
        if image_url == None:
            continue

        # Download and save the cover image
        # Consider sanitizing filename
        image_filename = os.path.join(
            COVER_IMAGES_PATH, f'{create_image_name(i+min_row, book_name, author_name)}.jpg')
        download_image(image_url, image_filename)

        sleep(1)


if __name__ == "__main__":
    main()
